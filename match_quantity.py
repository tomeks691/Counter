import openpyxl
import json
import string

items = {}
items_found = {}
sorted_item_found = {}
column = "B"
letters = string.ascii_uppercase


def get_parameters():
    while True:
        try:
            amount_sku = int(input("Podaj ilość SKU: "))
            break
        except:
            continue
    for _ in range(amount_sku):
        items[input("Podaj SKU: ")] = 0


def search_boxes():
    wb = openpyxl.load_workbook('test.xlsx')
    name_worksheet = input("Wprowadź nazwę arkusza który przeliczyć: ")
    ws = wb[name_worksheet]
    for i in range(1, ws.max_row + 1):
        if ws[f"{column}{i}"].value in items:
            if items_found.get(ws[f'B{i}'].value):
                items_found[ws[f"{column}{i}"].value] += [{ws[f"E{i}"].value: int(ws[f"D{i}"].value)}]
            else:
                items_found[ws[f"{column}{i}"].value] = [{ws[f"E{i}"].value: int(ws[f"D{i}"].value)}]

    return {k: sorted(v, key=lambda x: list(x.values())[0], reverse=True) for k, v in
            items_found.items()}


def appropriate_amount(sorted_items_found):
    item_amount_need = {}
    output_json = dict()
    while True:
        try:
            for item in items:
                item_amount = int(input(f"Podaj ilość do wyjęcia dla {item}: "))
                item_amount_need[item] = item_amount
            break
        except:
            continue
    for item in items:
        output_json[item] = []
        for item_found in sorted_items_found[item]:
            for box, amount in item_found.items():
                if item_amount_need[item] <= amount:
                    output_json[item] += [box, amount]
                    delete_amount(amount, item, box)
                    break
                else:
                    output_json[item] += [box, amount]
                    item_amount_need[item] -= amount
                    delete_amount(amount, item, box)
                    continue
            else:
                continue
            break
    with open("test.json", "w") as f:
        json.dump(output_json, f, indent=4)


def delete_amount(amount, sku, box):
    print(sku, box)
    wb = openpyxl.load_workbook('test.xlsx')
    ws = wb["ZK 47"]
    alphabet = string.ascii_uppercase
    for column in range(1, 6):
        for row in range(2, ws.max_row):
            if sku == ws[f"{alphabet[column]}{row}"].value and box == ws[f"{alphabet[column+3]}{row}"].value:
                amount_in_worksheet = ws[f"{alphabet[column+2]}{row}"].value
                amount_in_worksheet -= amount
                if amount_in_worksheet - amount >= 0:
                    ws.cell(row=row, column=4, value=amount_in_worksheet)
                else:
                    ws.cell(row=row, column=4, value=0)
    wb.save('PL.xlsx')

get_parameters()
output_json = appropriate_amount(search_boxes())

