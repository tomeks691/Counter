import openpyxl
import json

wb = openpyxl.load_workbook('test.xlsx')
name_worksheet = input("Wprowadź nazwę arkusza który przeliczyć: ")
ws = wb[name_worksheet]
count_shoes_in_pl = {}
for i in range(2, ws.max_row+1):
    if count_shoes_in_pl.get(ws[f'B{i}'].value):
        count_shoes_in_pl[ws[f'B{i}'].value] += ws[f'D{i}'].value
    else:
        count_shoes_in_pl[ws[f'B{i}'].value] = ws[f'D{i}'].value

print(count_shoes_in_pl)
with open("amount.json", "w") as f:
    json.dump(count_shoes_in_pl, f, indent=4)

